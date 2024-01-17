import pandas
import pandas as pd
import csv
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from operator import itemgetter
import time

#time right now (i.e at the time that the report is generated)
current_dateTime = time.localtime()
tymRytNow = time.strftime('%H:%M:%S', current_dateTime)
nowNow = datetime.strptime(tymRytNow, '%H:%M:%S').time()

#List of all students
with open('students.csv', mode='r') as csvfile:
    reader = csv.reader(csvfile)
    studentsList = []
    for row in reader:
        studentsList.append(row)


#convert xlsx to csv
file = pandas.read_excel("report.xlsx", converters={"User ID": str})
file.to_csv("report.csv", index=None, header=True)

#read csv data into records list
with open('report.csv', mode='r') as csvfile:
    reader = csv.reader(csvfile)
    records = []
    for row in reader:
        records.append([row[0], row[1], row[2]])

#remove header from list
records.pop(0)

#get the unique dates
uniqueDates = []
for x in records:
    datetimeObject = datetime.strptime(x[0], '%Y-%m-%d %H:%M:%S')
    date = datetimeObject.date().strftime("%Y-%m-%d")
    time = datetimeObject.time().strftime("%H:%M:%S")
    x[0] = date
    x.append(time)
    if date not in uniqueDates:
        uniqueDates.append(date)


#***Loop thru each date and make each date a key to e dict
mealsDict = {}
absentDict = {}
for u in uniqueDates:
    mealsDict[u] = {"breakfast": [],
                    "lunch": [],
                    "supper": [],}
    
    absentDict[u] = {"breakfast": [],
                    "lunch": [],
                    "supper": [],}
    
dictKeys = mealsDict.keys()


#nested loop 4 each record then each date, then append the record to its date and meal time
for r in records:
    date = datetime.strptime(r[0], '%Y-%m-%d').date()
    time = datetime.strptime(r[3], '%H:%M:%S').time()

    for u in uniqueDates:
        if u in r:
            if time > datetime.strptime('00:00:00', '%H:%M:%S').time() and time < datetime.strptime('09:00:00', '%H:%M:%S').time():
                mealsDict[u]["breakfast"] += [r]
            elif time > datetime.strptime('11:00:00', '%H:%M:%S').time() and time < datetime.strptime('15:00:00', '%H:%M:%S').time():
                mealsDict[u]["lunch"] += [r]
            elif time > datetime.strptime('16:00:00', '%H:%M:%S').time() and time < datetime.strptime('20:00:00', '%H:%M:%S').time():
                mealsDict[u]["supper"] += [r]


###Loop through the meal dict and find absent people###
tick = u'\u2713'
studentReport = {}
indiReport = []
for s in studentsList:
    mark = []
    count = 0

    #adding leading zeros to laundry numbas for keith mutandi, ewan burbidge & zara zietsman
    if s[0] == "590":
        s[0] = "0590"
    if s[0] == "873":
        s[0] = "0873"
    if s[0] == "92":
        s[0] = "092"
    
    for k in dictKeys:
        result =  any(s[0] in sublist for sublist in mealsDict[k]["breakfast"])
        if result == False:
            absentDict[k]["breakfast"] += s
            #mark.append(k + "_breakfast")
            mark.append("x")
            count = count + 1
        else:
            mark.append(tick)
        result =  any(s[0] in sublist for sublist in mealsDict[k]["lunch"])
        if result == False:
            absentDict[k]["lunch"] += s
            #mark.append(k + "_lunch")
            mark.append("x")
            count = count + 1
        else:
            mark.append(tick)
        result =  any(s[0] in sublist for sublist in mealsDict[k]["supper"])
        if result == False:
            absentDict[k]["supper"] += s
            #mark.append(k + "_supper")
            mark.append("x")
            count = count + 1
        else:
            mark.append(tick)    
    mark.reverse()

    #correct laundry numbers for students with incomplete laundry numbers
    if s[0] == "0590":
        s[0] = "O598"
    if s[0] == "0873":
        s[0] = "F873"
    if s[0] == "092":
        s[0] = "K092"
    if s[0] == "100":
        s[0] = "K608"
    if s[0] == "101":
        s[0] = "K094"
    if s[0] == "102":
        s[0] = "K081"
    if s[0] == "103":
        s[0] = "K047"
    if s[0] == "217":
        s[0] = "K217"
    if s[0] == "230":
        s[0] = "K230"
    if s[0] == "236":
        s[0] = "K236"
    if s[0] == "606":
        s[0] = "C606"
    if s[0] == "635":
        s[0] = "H635"
    if s[0] == "t825":
        s[0] = "T825"

    #indiReport.append([s[0], s[1], count, *mark])
    #use the current time to determine time of day, then subtract "count"
    if nowNow > datetime.strptime('00:00:00', '%H:%M:%S').time() and nowNow < datetime.strptime('11:59:00', '%H:%M:%S').time():
        indiReport.append([s[0], s[1], count-2, *mark])
    elif nowNow > datetime.strptime('12:00:00', '%H:%M:%S').time() and nowNow < datetime.strptime('15:59:00', '%H:%M:%S').time():
        indiReport.append([s[0], s[1], count-1, *mark])
    elif nowNow > datetime.strptime('16:00:00', '%H:%M:%S').time() and nowNow < datetime.strptime('23:59:00', '%H:%M:%S').time():
        indiReport.append([s[0], s[1], count, *mark])



#sorting the indieReport lexicographically by laundry number
sorted_indieReport = sorted(indiReport, key=itemgetter(0))


headers = []
for u in uniqueDates:
    headers.append(u + "_breakfast")
    headers.append(u + "_lunch")
    headers.append(u + "_supper")

headers.reverse()
headers.insert(0, "No. of Missed Meals")
headers.insert(0, "Name")
headers.insert(0, "Laundry No.")
print("HEADERS:")
print(len(headers))

#creating the dataframe and making it an excel
reportName = uniqueDates[-1] + "_Biometric Report.xlsx"
df = pd.DataFrame(sorted_indieReport)
df.to_excel(reportName, index=False, header=headers)

#open workbook and iterate through all cells
# load excel with its path 
wrkbk = openpyxl.load_workbook(reportName) 
  
sh = wrkbk.active 
  
# iterate through excel and display data
red = PatternFill(patternType='solid', fgColor="ff0000")
green = PatternFill(patternType='solid', fgColor="008000")
cellIds = []
for row in sh.iter_rows(min_row=0, min_col=1, max_row=len(studentsList)+1, max_col=len(uniqueDates*3)+3): 
    for cell in row: 
        cellIds.append(cell.coordinate) 

for i in range(len(cellIds)):
    if sh[cellIds[i]].value == "x":
        sh[cellIds[i]].fill = red
    elif sh[cellIds[i]].value == tick:
        sh[cellIds[i]].fill = green


wrkbk.save(reportName)



print(len(cellIds))

#with pd.ExcelWriter("indiReport.xlsx") as writer:
    #df.to_excel(writer)
